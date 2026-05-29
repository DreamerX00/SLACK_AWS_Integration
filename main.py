import os
import json
import sys
import time
from datetime import datetime
import logging
from typing import Dict, List, Optional, Any, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed

import boto3
import pandas as pd
from botocore.exceptions import ClientError, BotoCoreError
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

PRICING_REGION = "us-east-1"

AWS_REGION_MAP: Dict[str, str] = {
    "us-east-1": "US East (N. Virginia)",
    "us-east-2": "US East (Ohio)",
    "us-west-1": "US West (N. California)",
    "us-west-2": "US West (Oregon)",
    "ca-central-1": "Canada (Central)",
    "eu-west-1": "EU (Ireland)",
    "eu-west-2": "EU (London)",
    "eu-west-3": "EU (Paris)",
    "eu-central-1": "EU (Frankfurt)",
    "eu-central-2": "EU (Zurich)",
    "eu-north-1": "EU (Stockholm)",
    "eu-south-1": "EU (Milan)",
    "eu-south-2": "EU (Spain)",
    "ap-east-1": "Asia Pacific (Hong Kong)",
    "ap-south-1": "Asia Pacific (Mumbai)",
    "ap-south-2": "Asia Pacific (Hyderabad)",
    "ap-southeast-1": "Asia Pacific (Singapore)",
    "ap-southeast-2": "Asia Pacific (Sydney)",
    "ap-southeast-3": "Asia Pacific (Jakarta)",
    "ap-southeast-4": "Asia Pacific (Melbourne)",
    "ap-southeast-5": "Asia Pacific (Malaysia)",
    "ap-southeast-7": "Asia Pacific (Thailand)",
    "ap-northeast-1": "Asia Pacific (Tokyo)",
    "ap-northeast-2": "Asia Pacific (Seoul)",
    "ap-northeast-3": "Asia Pacific (Osaka)",
    "sa-east-1": "South America (Sao Paulo)",
    "me-south-1": "Middle East (Bahrain)",
    "me-central-1": "UAE (Dubai)",
    "af-south-1": "Africa (Cape Town)",
    "il-central-1": "Israel (Tel Aviv)",
}

REGION_DISPLAY: Dict[str, str] = {
    k: v.split("(")[-1].rstrip(")") if "(" in v else v
    for k, v in AWS_REGION_MAP.items()
}

EC2_PRICING_MODELS: List[Dict[str, Any]] = [
    {
        "label": "On-Demand",
        "source": "pricing",
        "term_type": "OnDemand",
        "lease_length": "",
        "purchase_option": "",
        "offering_class": "",
    },
    {
        "label": "EC2 SP 1yr No Upfront",
        "source": "savingsplans",
        "duration_years": 1,
        "payment_option": "No Upfront",
        "plan_type": "EC2Instance",
    },
    {
        "label": "EC2 SP 3yr No Upfront",
        "source": "savingsplans",
        "duration_years": 3,
        "payment_option": "No Upfront",
        "plan_type": "EC2Instance",
    },
    {
        "label": "Standard RI 1yr No Upfront",
        "source": "pricing",
        "term_type": "Reserved",
        "lease_length": "1yr",
        "purchase_option": "No Upfront",
        "offering_class": "standard",
    },
    {
        "label": "Standard RI 3yr No Upfront",
        "source": "pricing",
        "term_type": "Reserved",
        "lease_length": "3yr",
        "purchase_option": "No Upfront",
        "offering_class": "standard",
    },
]

RDS_PRICING_MODELS: List[Dict[str, Any]] = [
    {
        "label": "On-Demand",
        "source": "pricing",
        "term_type": "OnDemand",
        "lease_length": "",
        "purchase_option": "",
        "offering_class": "",
    },
    {
        "label": "Standard RI 1yr No Upfront",
        "source": "pricing",
        "term_type": "Reserved",
        "lease_length": "1yr",
        "purchase_option": "No Upfront",
        "offering_class": "standard",
    },
]

EC2_OS_NORMALIZE = {
    "linux": "Linux",
    "windows": "Windows",
    "rhel": "RHEL",
    "red hat enterprise linux": "RHEL",
    "suse": "SUSE",
    "suse linux": "SUSE",
}

SP_OS_MAP = {
    "Linux": "Linux/UNIX",
    "Windows": "Windows",
    "RHEL": "Red Hat Enterprise Linux",
    "SUSE": "SUSE Linux",
}

RdsEngineSpec = Tuple[str, Optional[str], Optional[str]]

RDS_ENGINE_SPEC: Dict[str, RdsEngineSpec] = {
    "mysql": ("MySQL", None, None),
    "postgresql": ("PostgreSQL", None, None),
    "postgres": ("PostgreSQL", None, None),
    "mariadb": ("MariaDB", None, None),
    "oracle": ("Oracle", "Standard Two", "Bring your own license"),
    "oracle se2": ("Oracle", "Standard Two", "Bring your own license"),
    "oracle se2 byol": ("Oracle", "Standard Two", "Bring your own license"),
    "oracle se2 li": ("Oracle", "Standard Two", "License included"),
    "oracle ee": ("Oracle", "Enterprise", "Bring your own license"),
    "oracle ee byol": ("Oracle", "Enterprise", "Bring your own license"),
    "sqlserver": ("SQL Server", "Standard", "License included"),
    "sql server": ("SQL Server", "Standard", "License included"),
    "mssql": ("SQL Server", "Standard", "License included"),
    "sql server se": ("SQL Server", "Standard", "License included"),
    "sql server se li": ("SQL Server", "Standard", "License included"),
    "sql server ee": ("SQL Server", "Enterprise", "License included"),
    "sql server ee li": ("SQL Server", "Enterprise", "License included"),
    "sql server web": ("SQL Server", "Web", "License included"),
    "sql server express": ("SQL Server", "Express", "License included"),
}


def normalize_ec2_os(os_name: str) -> str:
    return EC2_OS_NORMALIZE.get(os_name.strip().lower(), os_name)


def normalize_rds_engine(engine: str) -> RdsEngineSpec:
    return RDS_ENGINE_SPEC.get(engine.strip().lower(), (engine, None, None))

MAX_RETRIES = 3
RETRY_BASE_DELAY = 1.0

Ec2RowDef = Tuple[str, str, str]
RdsRowDef = Tuple[str, str, str]
SeedData = Tuple[List[Ec2RowDef], List[RdsRowDef]]


def _call_with_retry(client_method, **kwargs) -> Any:
    last_exc = None
    for attempt in range(MAX_RETRIES):
        try:
            return client_method(**kwargs)
        except Exception as e:
            last_exc = e
            if "Throttling" in str(e) or "Rate exceeded" in str(e):
                delay = RETRY_BASE_DELAY * (2 ** attempt)
                log.debug("Throttled, retrying in %.1fs (attempt %d/%d)", delay, attempt + 1, MAX_RETRIES)
                time.sleep(delay)
            else:
                raise
    raise last_exc


def get_aws_clients() -> Tuple[boto3.client, boto3.client]:
    aws_access_key = os.getenv("AWS_ACCESS_KEY_ID")
    aws_secret_key = os.getenv("AWS_SECRET_ACCESS_KEY")
    aws_session_token = os.getenv("AWS_SESSION_TOKEN")

    session_kwargs: Dict[str, Any] = {}
    if aws_access_key and aws_secret_key:
        session_kwargs["aws_access_key_id"] = aws_access_key
        session_kwargs["aws_secret_access_key"] = aws_secret_key
    if aws_session_token:
        session_kwargs["aws_session_token"] = aws_session_token
    if not session_kwargs:
        log.info("AWS credentials not in .env; using default credential chain.")

    pricing = boto3.client("pricing", region_name=PRICING_REGION, **session_kwargs)
    savings = boto3.client("savingsplans", region_name=PRICING_REGION, **session_kwargs)
    return pricing, savings


def extract_price_from_terms(
    terms: Dict[str, Any],
    term_type: str,
    lease_length: str,
    purchase_option: str,
    offering_class: str = "standard",
) -> Optional[float]:
    term_data = terms.get(term_type)
    if not term_data:
        return None

    for sku_key, term_entry in term_data.items():
        if not isinstance(term_entry, dict):
            continue

        ta = term_entry.get("termAttributes", {})

        if term_type == "Reserved":
            if (
                ta.get("LeaseContractLength", "") != lease_length
                or ta.get("PurchaseOption", "") != purchase_option
                or (offering_class and ta.get("OfferingClass", "") != offering_class)
            ):
                continue

        price_dimensions = term_entry.get("priceDimensions", {})
        if not isinstance(price_dimensions, dict):
            continue

        hourly_rate = None
        for price_dim in price_dimensions.values():
            if not isinstance(price_dim, dict):
                continue
            unit = price_dim.get("unit", "")
            if unit != "Hrs":
                continue
            price_per_unit = price_dim.get("pricePerUnit", {}).get("USD")
            if price_per_unit is None:
                continue
            try:
                val = float(price_per_unit)
            except (ValueError, TypeError):
                continue
            hourly_rate = val
            break

        if hourly_rate is not None:
            return hourly_rate

    return None


def fetch_ec2_pricing_api(
    client: boto3.client,
    instance_type: str,
    location: str,
    os_name: str,
    model: Dict[str, Any],
) -> Optional[float]:
    normalized_os = normalize_ec2_os(os_name)

    filters = [
        {"Type": "TERM_MATCH", "Field": "instanceType", "Value": instance_type},
        {"Type": "TERM_MATCH", "Field": "location", "Value": location},
        {"Type": "TERM_MATCH", "Field": "operatingSystem", "Value": normalized_os},
        {"Type": "TERM_MATCH", "Field": "tenancy", "Value": "Shared"},
        {"Type": "TERM_MATCH", "Field": "preInstalledSw", "Value": "NA"},
        {"Type": "TERM_MATCH", "Field": "capacityStatus", "Value": "Used"},
        {"Type": "TERM_MATCH", "Field": "licenseModel", "Value": "No License required"},
    ]

    try:
        response = _call_with_retry(
            client.get_products, ServiceCode="AmazonEC2", Filters=filters, MaxResults=100
        )
    except Exception as e:
        log.warning("Pricing API error (%s %s %s): %s", instance_type, location, os_name, e)
        return None

    best_price = None
    for price_list_str in response.get("PriceList", []):
        try:
            price_data = json.loads(price_list_str)
        except json.JSONDecodeError:
            continue
        price = extract_price_from_terms(
            price_data.get("terms", {}),
            model["term_type"],
            model["lease_length"],
            model["purchase_option"],
            model["offering_class"],
        )
        if price is not None:
            if best_price is None or price < best_price:
                best_price = price

    return best_price


def fetch_ec2_sp_api(
    client: boto3.client,
    instance_type: str,
    region: str,
    os_name: str,
    model: Dict[str, Any],
) -> Optional[float]:
    normalized_os = normalize_ec2_os(os_name)
    sp_os = SP_OS_MAP.get(normalized_os)
    if not sp_os:
        log.warning("Unknown OS '%s' for SP API lookup; skipping.", os_name)
        return None

    filters = [
        {"name": "region", "values": [region]},
        {"name": "instanceType", "values": [instance_type]},
        {"name": "tenancy", "values": ["shared"]},
        {"name": "productDescription", "values": [sp_os]},
    ]

    try:
        response = _call_with_retry(
            client.describe_savings_plans_offering_rates,
            savingsPlanTypes=[model["plan_type"]],
            savingsPlanPaymentOptions=[model["payment_option"]],
            products=["EC2"],
            filters=filters,
        )
        duration_target = model["duration_years"] * 31536000
        matching = []
        for rate_info in response.get("searchResults", []):
            sp_offer = rate_info.get("savingsPlanOffering", {})
            if sp_offer.get("durationSeconds") != duration_target:
                continue
            unit = rate_info.get("unit", "")
            if unit and unit != "Hrs":
                continue
            try:
                rate_val = float(rate_info.get("rate", 0))
            except (ValueError, TypeError):
                continue
            if rate_val > 0:
                matching.append(rate_val)
        if matching:
            return min(matching)
    except Exception as e:
        log.warning("SP API error (%s %s %s): %s", instance_type, region, os_name, e)

    return None


def fetch_rds_price(
    client: boto3.client,
    instance_type: str,
    location: str,
    database_engine: str,
    model: Dict[str, Any],
) -> Optional[float]:
    engine, edition, license_model = normalize_rds_engine(database_engine)

    filters = [
        {"Type": "TERM_MATCH", "Field": "instanceType", "Value": instance_type},
        {"Type": "TERM_MATCH", "Field": "location", "Value": location},
        {"Type": "TERM_MATCH", "Field": "databaseEngine", "Value": engine},
        {"Type": "TERM_MATCH", "Field": "deploymentOption", "Value": "Single-AZ"},
    ]
    if edition:
        filters.append({"Type": "TERM_MATCH", "Field": "databaseEdition", "Value": edition})
    if license_model:
        filters.append({"Type": "TERM_MATCH", "Field": "licenseModel", "Value": license_model})

    try:
        response = _call_with_retry(
            client.get_products, ServiceCode="AmazonRDS", Filters=filters, MaxResults=100
        )
    except Exception as e:
        log.warning("RDS API error (%s %s %s): %s", instance_type, location, database_engine, e)
        return None

    candidate_prices: List[float] = []
    for price_list_str in response.get("PriceList", []):
        try:
            price_data = json.loads(price_list_str)
        except json.JSONDecodeError:
            continue
        attrs = price_data.get("product", {}).get("attributes", {})
        # InstanceUsage SKUs only — exclude HeavyUsage/ReservedUsage etc.
        # Region prefix varies: "InstanceUsage:db.t3.micro" in us-east-1,
        # "APS3-InstanceUsage:db.t3.micro" in ap-south-1, etc.
        usagetype = attrs.get("usagetype", "")
        if "InstanceUsage" not in usagetype:
            continue
        price = extract_price_from_terms(
            price_data.get("terms", {}),
            model["term_type"],
            model["lease_length"],
            model["purchase_option"],
            model["offering_class"],
        )
        if price is not None and price > 0:
            candidate_prices.append(price)

    if not candidate_prices:
        return None
    return min(candidate_prices)


def _fetch_ec2_row(
    pricing_client: boto3.client,
    sp_client: boto3.client,
    instance_type: str,
    region: str,
    location: str,
    os_name: str,
) -> Optional[Dict[str, Any]]:
    region_display = REGION_DISPLAY.get(region, region)
    normalized_os = normalize_ec2_os(os_name)
    row: Dict[str, Any] = {
        "Service": "EC2",
        "Region": region_display,
        "Instance Type": instance_type,
        "OS": normalized_os,
    }
    has_any = False

    for model in EC2_PRICING_MODELS:
        if model["source"] == "savingsplans":
            price = fetch_ec2_sp_api(sp_client, instance_type, region, normalized_os, model)
        else:
            price = fetch_ec2_pricing_api(pricing_client, instance_type, location, normalized_os, model)
        row[model["label"]] = price if price is not None else "N/A"
        if price is not None:
            has_any = True

    return row if has_any else None


def _fetch_rds_row(
    client: boto3.client,
    instance_type: str,
    region: str,
    location: str,
    engine: str,
) -> Optional[Dict[str, Any]]:
    region_display = REGION_DISPLAY.get(region, region)
    eng_name, edition, license_model = normalize_rds_engine(engine)
    display_engine = eng_name
    if edition:
        display_engine = f"{eng_name} {edition}"
    if license_model:
        license_tag = "BYOL" if license_model.lower().startswith("bring") else "LI"
        display_engine = f"{display_engine} ({license_tag})"
    row: Dict[str, Any] = {
        "Service": "RDS",
        "Region": region_display,
        "Instance Type": instance_type,
        "Engine": display_engine,
    }
    has_any = False

    for model in RDS_PRICING_MODELS:
        price = fetch_rds_price(client, instance_type, location, engine, model)
        row[model["label"]] = price if price is not None else "N/A"
        if price is not None:
            has_any = True

    return row if has_any else None


def build_ec2_rows(
    pricing_client: boto3.client,
    sp_client: boto3.client,
    ec2_tuples: List[Ec2RowDef],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    tasks = {}

    with ThreadPoolExecutor(max_workers=10) as executor:
        for idx, (instance_type, region, os_name) in enumerate(ec2_tuples):
            location = AWS_REGION_MAP.get(region)
            if not location:
                log.warning("Unknown region mapping: %s. Skipping.", region)
                continue
            fut = executor.submit(_fetch_ec2_row, pricing_client, sp_client, instance_type, region, location, os_name)
            tasks[fut] = idx

        for future in as_completed(tasks):
            try:
                row = future.result()
                if row:
                    row["_sort_idx"] = tasks[future]
                    rows.append(row)
            except Exception as e:
                log.error("EC2 task failed: %s", e)

    rows.sort(key=lambda r: r.pop("_sort_idx"))
    return rows


def build_rds_rows(
    client: boto3.client,
    rds_tuples: List[RdsRowDef],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    tasks = {}

    with ThreadPoolExecutor(max_workers=10) as executor:
        for idx, (instance_type, region, engine) in enumerate(rds_tuples):
            location = AWS_REGION_MAP.get(region)
            if not location:
                log.warning("Unknown region mapping: %s. Skipping.", region)
                continue
            fut = executor.submit(_fetch_rds_row, client, instance_type, region, location, engine)
            tasks[fut] = idx

        for future in as_completed(tasks):
            try:
                row = future.result()
                if row:
                    row["_sort_idx"] = tasks[future]
                    rows.append(row)
            except Exception as e:
                log.error("RDS task failed: %s", e)

    rows.sort(key=lambda r: r.pop("_sort_idx"))
    return rows


def write_excel_report(
    ec2_rows: List[Dict[str, Any]],
    rds_rows: List[Dict[str, Any]],
    output_path: str = "",
) -> str:
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"aws_cost_report_{ts}.xlsx"
    df_ec2 = pd.DataFrame(ec2_rows) if ec2_rows else pd.DataFrame()
    df_rds = pd.DataFrame(rds_rows) if rds_rows else pd.DataFrame()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not df_ec2.empty:
            df_ec2.to_excel(writer, sheet_name="EC2 Pricing", index=False)
        if not df_rds.empty:
            df_rds.to_excel(writer, sheet_name="RDS Pricing", index=False)
        if df_ec2.empty and df_rds.empty:
            pd.DataFrame({"Status": ["No data fetched. Check inputs and AWS credentials."]}).to_excel(
                writer, sheet_name="Summary", index=False
            )

    _format_excel_sheets(output_path)
    log.info("Report saved to %s", output_path)
    return output_path


def _format_excel_sheets(filepath: str) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(filepath)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        na_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for ws in wb.worksheets:
            ws.freeze_panes = "B2"

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if isinstance(cell.value, float):
                        cell.number_format = "$#,##0.0000"
                    elif isinstance(cell.value, str) and cell.value == "N/A":
                        cell.fill = na_fill

            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    (
                        len(str(cell.value))
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx)
                        for cell in row
                        if cell.value
                    ),
                    default=8,
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

        wb.save(filepath)

    except ImportError:
        log.warning("openpyxl formatting not available; skipping Excel styling.")
    except Exception as e:
        log.warning("Excel formatting error (non-fatal): %s", e)


VALID_OS_OPTIONS = ["Linux", "Windows", "RHEL", "SUSE"]
VALID_RDS_ENGINES = [
    "MySQL", "PostgreSQL", "MariaDB",
    "Oracle SE2 BYOL", "Oracle SE2 LI", "Oracle EE BYOL",
    "SQL Server SE LI", "SQL Server EE LI", "SQL Server Web", "SQL Server Express",
]
VALID_REGIONS = list(AWS_REGION_MAP.keys())


def _prompt_list(prompt_text: str, valid_options: Optional[List[str]] = None) -> List[str]:
    raw = input(prompt_text).strip()
    items = [x.strip() for x in raw.split(",") if x.strip()]
    if not items:
        return []
    if valid_options:
        valid_lower = {v.lower(): v for v in valid_options}
        normalized = []
        for item in items:
            match = valid_lower.get(item.lower())
            if match:
                normalized.append(match)
            else:
                print(f"  Warning: '{item}' is not in the valid list. Proceeding anyway.")
                normalized.append(item)
        return normalized
    return items


def load_seed(path: str) -> SeedData:
    with open(path) as f:
        cfg = json.load(f)
    ec2 = cfg.get("ec2", {})
    rds = cfg.get("rds", {})

    if "rows" in ec2:
        ec2_tuples = [(r["instance_type"], r["region"], r["os"]) for r in ec2["rows"]]
    else:
        its = ec2.get("instance_types", ["t3.micro"])
        regs = ec2.get("regions", ["us-east-1"])
        oss = ec2.get("os_list", ["Linux"])
        ec2_tuples = [(it, reg, os) for it in its for reg in regs for os in oss]

    if "rows" in rds:
        rds_tuples = [(r["instance_type"], r["region"], r["engine"]) for r in rds["rows"]]
    else:
        its = rds.get("instance_types", ["db.t3.micro"])
        regs = rds.get("regions", ["us-east-1"])
        engs = rds.get("engines", ["MySQL"])
        rds_tuples = [(it, reg, eng) for it in its for reg in regs for eng in engs]

    return ec2_tuples, rds_tuples


def main():
    log.info("=" * 60)
    log.info("AWS Cost Report Generator (EC2 SP + RI + On-Demand)")
    log.info("=" * 60)

    try:
        pricing_client, sp_client = get_aws_clients()
    except Exception as e:
        log.error("Failed to initialize AWS clients: %s", e)
        sys.exit(1)

    if len(sys.argv) > 1 and sys.argv[1] == "--seed":
        if len(sys.argv) < 3:
            log.error("Usage: python main.py --seed <file.json>")
            sys.exit(1)
        ec2_tuples, rds_tuples = load_seed(sys.argv[2])
        log.info("Loaded seed from %s (%d EC2 rows, %d RDS rows)", sys.argv[2], len(ec2_tuples), len(rds_tuples))
    else:
        print("\n--- EC2 Input ---")
        ec2_it = _prompt_list("Instance types (comma-separated, e.g. t3.micro,m5.large): ") or ["t3.micro"]
        print(f"Valid regions: {', '.join(VALID_REGIONS)}")
        ec2_reg = _prompt_list("Regions (comma-separated, e.g. us-east-1,ap-south-1): ") or ["us-east-1"]
        print(f"Valid OS: {', '.join(VALID_OS_OPTIONS)}")
        ec2_os = _prompt_list("OS (comma-separated, e.g. Linux,Windows): ") or ["Linux"]
        ec2_tuples = [(it, reg, os) for it in ec2_it for reg in ec2_reg for os in ec2_os]

        print("\n--- RDS Input ---")
        rds_it = _prompt_list("Instance types (comma-separated, e.g. db.t3.micro,db.m5.large): ") or ["db.t3.micro"]
        print(f"Valid regions: {', '.join(VALID_REGIONS)}")
        rds_reg = _prompt_list("Regions (comma-separated, e.g. us-east-1,eu-west-1): ") or ["us-east-1"]
        print(f"Valid engines: {', '.join(VALID_RDS_ENGINES)}")
        rds_eng = _prompt_list("Engines (comma-separated, e.g. MySQL,PostgreSQL): ") or ["MySQL"]
        rds_tuples = [(it, reg, eng) for it in rds_it for reg in rds_reg for eng in rds_eng]

    log.info("Total pricing lookups: %d", len(ec2_tuples) + len(rds_tuples))

    ec2_rows = build_ec2_rows(pricing_client, sp_client, ec2_tuples) if ec2_tuples else []
    rds_rows = build_rds_rows(pricing_client, rds_tuples) if rds_tuples else []

    if not ec2_rows and not rds_rows:
        log.error("No pricing data fetched. Check inputs and AWS credentials.")
        sys.exit(1)

    output_path = write_excel_report(ec2_rows, rds_rows)
    print(f"\nReport generated: {os.path.abspath(output_path)}")
    print(f"  EC2 rows: {len(ec2_rows)}")
    print(f"  RDS rows: {len(rds_rows)}")


if __name__ == "__main__":
    main()
