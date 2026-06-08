"""Bridge between the Slack bot and the existing AWS pricing logic."""

from typing import Any, Dict

from openpyxl import load_workbook

from main import get_aws_clients, build_ec2_rows, build_rds_rows, write_excel_report

REQUIRED_EC2_COLS = {"instance_type", "region", "os"}
REQUIRED_RDS_COLS = {"instance_type", "region", "engine"}

SheetData = Dict[str, Any]


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_text(value: Any) -> str:
    return str(value).strip()


def _read_sheet_data(worksheet) -> SheetData:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [
        _normalize_text(cell_value).lower() if not _is_blank(cell_value) else ""
        for cell_value in header_row
    ]

    rows = []
    non_empty_row_count = 0

    for row_values in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(not _is_blank(cell_value) for cell_value in row_values):
            continue

        non_empty_row_count += 1
        row = {}
        for index, header in enumerate(headers):
            if header:
                row[header] = row_values[index] if index < len(row_values) else None
        rows.append(row)

    return {
        "columns": {header for header in headers if header},
        "rows": rows,
        "row_count": non_empty_row_count,
    }


def parse_input_workbook(input_path: str) -> Dict[str, SheetData]:
    workbook = load_workbook(filename=input_path, data_only=True, read_only=True)
    try:
        return {
            worksheet.title: _read_sheet_data(worksheet)
            for worksheet in workbook.worksheets
        }
    finally:
        workbook.close()


def _sheet_matches(sheet_name: str, sheet_data: SheetData, required_columns: set[str]) -> bool:
    """A sheet is treated as EC2/RDS if it is named accordingly OR its columns match.

    This makes the bot tolerant of workbooks where the data lives in a default-named
    sheet (e.g. "Sheet1") instead of a sheet titled exactly "EC2"/"RDS".
    """
    return required_columns.issubset(sheet_data["columns"])


def _classify_sheets(sheets: Dict[str, SheetData]) -> tuple[SheetData | None, SheetData | None]:
    """Return (ec2_sheet_data, rds_sheet_data) detected by columns, with name as a tiebreaker."""
    ec2_sheet: SheetData | None = None
    rds_sheet: SheetData | None = None

    for sheet_name, sheet_data in sheets.items():
        name_upper = sheet_name.strip().upper()
        is_ec2 = _sheet_matches(sheet_name, sheet_data, REQUIRED_EC2_COLS)
        is_rds = _sheet_matches(sheet_name, sheet_data, REQUIRED_RDS_COLS)

        # If a sheet's columns satisfy both (shouldn't normally happen), use the name to decide.
        if is_ec2 and is_rds:
            if name_upper == "RDS":
                is_ec2 = False
            else:
                is_rds = False

        if is_ec2 and ec2_sheet is None:
            ec2_sheet = sheet_data
        elif is_rds and rds_sheet is None:
            rds_sheet = sheet_data

    return ec2_sheet, rds_sheet


def _count_valid_rows(sheet_data: SheetData | None, required_columns: set[str]) -> int:
    if not sheet_data:
        return 0

    if not required_columns.issubset(sheet_data["columns"]):
        return 0

    return sum(
        1
        for row in sheet_data["rows"]
        if all(not _is_blank(row.get(column_name)) for column_name in required_columns)
    )


def inspect_input_workbook(input_path: str) -> Dict[str, int]:
    sheets = parse_input_workbook(input_path)
    ec2_sheet, rds_sheet = _classify_sheets(sheets)
    return {
        "total_rows": sum(sheet_data["row_count"] for sheet_data in sheets.values()),
        "ec2_count": _count_valid_rows(ec2_sheet, REQUIRED_EC2_COLS),
        "rds_count": _count_valid_rows(rds_sheet, REQUIRED_RDS_COLS),
    }


def _extract_service_tuples(
    sheet_data: SheetData,
    required_columns: set[str],
    field_order: tuple[str, ...],
    sheet_name: str,
) -> list[tuple[str, ...]]:
    lower_cols = sheet_data["columns"]
    if not required_columns.issubset(lower_cols):
        raise ValueError(
            f"{sheet_name} sheet must contain columns: {', '.join(sorted(required_columns))}. "
            f"Got: {', '.join(sorted(lower_cols))}"
        )

    tuples = []
    for row in sheet_data["rows"]:
        values = [row.get(field_name) for field_name in field_order]
        if any(_is_blank(value) for value in values):
            continue
        tuples.append(tuple(_normalize_text(value) for value in values))
    return tuples


def generate_cost_report(input_path: str, output_path: str) -> str:
    sheets = parse_input_workbook(input_path)
    ec2_sheet, rds_sheet = _classify_sheets(sheets)

    ec2_tuples = []
    rds_tuples = []

    if ec2_sheet is not None:
        ec2_tuples = _extract_service_tuples(
            ec2_sheet,
            REQUIRED_EC2_COLS,
            ("instance_type", "region", "os"),
            "EC2",
        )

    if rds_sheet is not None:
        rds_tuples = _extract_service_tuples(
            rds_sheet,
            REQUIRED_RDS_COLS,
            ("instance_type", "region", "engine"),
            "RDS",
        )

    if not ec2_tuples and not rds_tuples:
        raise ValueError(
            "Input file must contain EC2 columns (instance_type, region, os) "
            "and/or RDS columns (instance_type, region, engine)."
        )

    pricing_client, sp_client = get_aws_clients()

    ec2_rows = build_ec2_rows(pricing_client, sp_client, ec2_tuples) if ec2_tuples else []
    rds_rows = build_rds_rows(pricing_client, rds_tuples) if rds_tuples else []

    if not ec2_rows and not rds_rows:
        raise RuntimeError("No pricing data could be fetched. Check AWS credentials and inputs.")

    return write_excel_report(ec2_rows, rds_rows, output_path)
